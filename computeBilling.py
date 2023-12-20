import openpyxl
import requests
import json

# Open the Excel file
workbook = openpyxl.load_workbook('VM.xlsx')

# Select the sheet you want to work with
sheet = workbook['Sheet1']

# A list
column = []
# Extracting values to a list
for row in sheet.iter_rows(min_row=2, max_col=1, max_row=sheet.max_row):
    column.append(row[0].value)

# Print the list
print("Machine types [LIST] : ")
print(column)

print(" COST : ")
# Load the Excel file
# df = pd.read_excel('VM.xlsx')
# data = df.to_dict(orient='dict')
# print(data)

body = json.dumps({
    "costScenario": {
        "scenarioConfig": {
            "estimateDuration": "2628000s"
        },
        "workloads": [
            {
                "name": "vm-example",
                "computeVmWorkload": {
                    "instancesRunning": {
                        "usageRateTimeline": {
                            "usageRateTimelineEntries": [
                                {
                                    "usageRate": 1
                                }
                            ]
                        }
                    },
                    "machineType": {
                        "predefinedMachineType": {
                            "machineType": "e2-micro"
                        }
                    },
                    "region": "us-central1"
                }
            }
        ]
    }
})

URL = "https://cloudbilling.googleapis.com/v1beta:estimateCostScenario"
# f = {'file': open('VM.xlsx', 'rb')}
headers = {
    "Content-Type": "application/json",
    "Accept": "*/*",
    "Connection": "keep-alive",
    "X-goog-api-key": "AIzaSyAIdlAvVrTJbEkVCuv-WZ76A6uFMRTl_ZU"
}

response = requests.post(URL, headers=headers, data=body)
estimate = json.loads(response.text)["costEstimationResult"]['segmentCostEstimates'][0]["segmentTotalCostEstimate"][
    "preCreditCostEstimate"]
price = float(estimate["units"] + "." + str(estimate["nanos"]))
print(price,
      json.loads(response.text)["costEstimationResult"]['segmentCostEstimates'][0]["segmentTotalCostEstimate"][
          "preCreditCostEstimate"]["currencyCode"])

