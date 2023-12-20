import openpyxl
import requests
import json

# Open the Excel file
workbook = openpyxl.load_workbook('cudVM.xlsx')

# Select the sheet you want to work with
sheet = workbook['Sheet1']

for row in range(2, sheet.max_row + 1):
    # for cell in row:
    noOfInstances = sheet.cell(row=row, column=1).value
    location = sheet.cell(row=row, column=2).value
    machineType = sheet.cell(row=row, column=3).value
    vCpuCount = sheet.cell(row=row, column=4).value
    machineSeries = str(machineType)
    machineSeries = machineType.split("-")[0]
    memSize = sheet.cell(row=row, column=5).value
    plan = sheet.cell(row=row, column=6).value
    gpu = sheet.cell(row=row, column=7).value
    gpuTotal = sheet.cell(row=row, column=8).value

    body = json.dumps(
        {
            "costScenario": {
                "workloads": [{
                    "name": "CUDDiscounts",
                    "computeVmWorkload": {
                        "region": location,
                        "machineType": {
                            "predefinedMachineType": {
                                "machineType": machineType
                            }
                        },
                        "instancesRunning": {
                            "usageRateTimeline": {
                                "usageRateTimelineEntries": [{
                                    "usageRate": 1
                                }]
                            }
                        }
                    }
                }],
                "commitments": [{
                    "name": "CUDDetails",
                    "vmResourceBasedCud": {
                        "region": location,
                        "virtualCpuCount": vCpuCount,
                        "memorySizeGb": memSize,
                        "plan": plan,
                        "machineSeries": machineSeries
                    }
                }],
                "scenarioConfig": {
                    "estimateDuration": "2628000s"
                }
            }
        }
    )

    URL = "https://cloudbilling.googleapis.com/v1beta:estimateCostScenario"
    
    # Headers - Sensitive Info - DICTIONARY
    headers = {
        "Content-Type": "application/json",
        "Accept": "*/*",
        "Connection": "keep-alive",
        "X-goog-api-key": "AIzaSyAIdlAvVrTJbEkVCuv-WZ76A6uFMRTl_ZU"
    }

    response = requests.post(URL, headers=headers, data=body)
    estimate = json.loads(response.text)["costEstimationResult"]['segmentCostEstimates'][0]["segmentTotalCostEstimate"][
        "netCostEstimate"]
    print(estimate)

    if len(estimate) == 2:
        price = str(estimate["units"] + ".00")
        sheet.cell(row=row, column=7).value = round((float(price)), 2)
    else:
        if len(str(estimate["nanos"])) == 9:
            price = float(estimate["units"] + "." + str(estimate["nanos"]))
            sheet.cell(row=row, column=7).value = (round(price, 2))
        else:
            price = float(estimate["units"] + "." + "0" * (9 - (len(str(estimate["nanos"])))) + str(estimate["nanos"]))
            sheet.cell(row=row, column=7).value = (round(price, 2))
workbook.save('cudVM.xlsx')
