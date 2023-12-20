import openpyxl
import requests
import json

# Open the Excel file
workbook = openpyxl.load_workbook('premiumVM.xlsx')

# Select the sheet you want to work with
sheet = workbook['Sheet1']

for row in range(2, sheet.max_row + 1):
    instance_type = str(sheet.cell(row=row, column=1).value)
    machineSeries = str(instance_type).split("-", maxsplit=1)
    machineSeries = machineSeries[0]
    region = str(sheet.cell(row=row, column=2).value)
    vCpu = int(sheet.cell(row=row, column=3).value)
    memGB = float(sheet.cell(row=row, column=4).value)
    body = json.dumps(
        {
            "costScenario": {
                "scenarioConfig": {
                    "estimateDuration": "2592000s"
                },
                "workloads": [
                    {
                        "name": "vm-example",
                        "computeVmWorkload": {
                            "instancesRunning": {
                                "usageRateTimeline": {
                                    "usageRateTimelineEntries": [
                                        {
                                            "usageRate": 1
                                        }
                                    ]
                                }
                            },
                            "machineType": {
                                "customMachineType": {
                                    "machineSeries": machineSeries,
                                    "virtualCpuCount": vCpu,
                                    "memorySizeGb": memGB
                                }
                            },
                            "region": region,
                            "licenses": "projects/windows-cloud/global/licenses/windows-server-2022-dc"
                        }
                    }
                ]
            }
        }
    )
    URL = "https://cloudbilling.googleapis.com/v1beta:estimateCostScenario"

    headers = {
        "Content-Type": "application/json",
        "Accept": "*/*",
        "Connection": "keep-alive",
        "X-goog-api-key": "AIzaSyAIdlAvVrTJbEkVCuv-WZ76A6uFMRTl_ZU"
    }

    response = requests.post(URL, headers=headers, data=body)

    estimate = json.loads(response.text)["costEstimationResult"]['segmentCostEstimates'][0]["segmentTotalCostEstimate"][
        "preCreditCostEstimate"]
    print (estimate)
    if len(estimate) == 2:
        price = str(estimate["units"] + ".00")
        sheet.cell(row=row, column=5).value = round((float(price)), 2)
    else:
        if len(str(estimate["nanos"])) == 9:
            price = float(estimate["units"] + "." + str(estimate["nanos"]))
            sheet.cell(row=row, column=5).value = (round(price, 2))
        else:
            price = float(estimate["units"] + "." + "0" * (9 - (len(str(estimate["nanos"])))) + str(estimate["nanos"]))
            sheet.cell(row=row, column=5).value = (round(price, 2))
workbook.save('premiumVM.xlsx')
