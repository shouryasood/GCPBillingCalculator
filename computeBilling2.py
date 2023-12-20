import openpyxl
import requests
import json

# Open the Excel file
workbook = openpyxl.load_workbook('VM.xlsx')

# Select the sheet you want to work with
sheet = workbook['Sheet1']

for row in range(2, sheet.max_row + 1):
    instance_type = str(sheet.cell(row=row, column=1).value)
    region = str(sheet.cell(row=row, column=2).value)
    diskType = str(sheet.cell(row=row, column=3).value)
    usage = (sheet.cell(row=row, column=4).value)# Usage of PD in [GiB]
    eIOPS= (sheet.cell(row=row, column=5).value) #Extreme PD IOPS [>2500]

    body = json.dumps(
        {
            "costScenario": {
                "scenarioConfig": {
                    "estimateDuration": "360000s"
                },
                "workloads": [
                    {
                        "name": "vm-example",
                        "computeVmWorkload": {
                            "instancesRunning": {
                                "usageRateTimeline": {
                                    "usageRateTimelineEntries": [
                                        {
                                            "usageRate": 5
                                        }
                                    ]
                                }
                            },
                            "machineType": {
                                "predefinedMachineType":
                                    {
                                        "machineType": instance_type
                                    }
                            },
                            "region": region,
                            "persistentDisks": [
                                {
                                        "diskType": diskType,
                                        "scope": "SCOPE_UNSPECIFIED",
                                        "diskSize": {
                                            "usageRateTimeline": {
                                                "unit": "GiBy",
                                                "usageRateTimelineEntries": [
                                                    {
                                                        "effectiveTime": {
                                                            "estimationTimeFrameOffset": "360000s"
                                                        },
                                                        "usageRate": usage
                                                    }
                                                ]
                                            }
                                        },
                                        "provisionedIops": {
                                            "usageRateTimeline": {
                                                "unit": "1/s",
                                                "usageRateTimelineEntries": [
                                                    {
                                                        "effectiveTime": {
                                                            "estimationTimeFrameOffset": "360000s"
                                                        },
                                                        "usageRate": eIOPS
                                                    }
                                                ]
                                            }
                                        }
                                }
                            ]
                        }
                    }
                ]
            }
        }
    )
    URL = "https://cloudbilling.googleapis.com/v1beta:estimateCostScenario"

    headers = {
        "Content-Type": "application/json",
        "Accept": "*/*",
        "Connection": "keep-alive",
        "X-goog-api-key": "AIzaSyAIdlAvVrTJbEkVCuv-WZ76A6uFMRTl_ZU"
    }

    response = requests.post(URL, headers=headers, data=body)

    estimate = json.loads(response.text)["costEstimationResult"]['segmentCostEstimates'][0]["segmentTotalCostEstimate"][
        "preCreditCostEstimate"]
    print (estimate)
    if len(estimate) == 2:
        price = str(estimate["units"] + ".00")
        sheet.cell(row=row, column=6).value = round((float(price)), 2)
    else:
        if len(str(estimate["nanos"])) == 9:
            price = float(estimate["units"] + "." + str(estimate["nanos"]))
            sheet.cell(row=row, column=6).value = (round(price, 2))
        else:
            price = float(estimate["units"] + "." + "0" * (9 - (len(str(estimate["nanos"])))) + str(estimate["nanos"]))
            sheet.cell(row=row, column=6).value = (round(price, 2))
workbook.save('VM.xlsx')
