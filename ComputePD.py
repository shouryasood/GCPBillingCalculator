import openpyxl
import requests
import json

# Open the Excel file
workbook = openpyxl.load_workbook('ComputePD.xlsx')

# Select the sheet you want to work with
sheet = workbook['Sheet1']

for row in range(2, sheet.max_row + 1):

    #Zonal Standard PD [GiB], Zonal SSD PD [GiB], Zonal Balanced PD [GiB], Extreme PD [GiB]
    diskType = str(sheet.cell(row=row, column=1).value)
    usage = (sheet.cell(row=row, column=2).value)# Usage of PD in [GiB]
    eIOPS= (sheet.cell(row=row, column=3).value) #Extreme PD IOPS [>2500]
    location = (sheet.cell(row=row, column=4).value) #Region where storage is located.
    scope = ["SCOPE_UNSPECIFIED","SCOPE_ZONAL", "SCOPE_REGIONAL"]

    
    body = json.dumps(
        {
            "costScenario": {
                "scenarioConfig": {
                    "estimateDuration": "2592000s"
                },
                
                "workloads": [
                    {
                        "name": "demoWorkload",
                        "computeVmWorkload": {
                            "machineType": {
                                
                            },
                            "guestAccelerator": {
                                
                            },
                            "preemptible": "true",
                            "enableConfidentialCompute": "false",
                            "licenses": [
                                "projects/centos-cloud/global/licenses/centos-stream-9"
                            ],
                            "persistentDisks": 
                            [
                                {
                                    "diskType": diskType,
                                    "scope": {
                                        "usageRateTimeline": {
                                            "unit": "GiBy",
                                            "usageRateTimelineEntries": [
                                                {
                                                    "estimationTimeFrameOffset":  "SCOPE_UNSPECIFIED"
                                                }
                                            ]
                                        }
                                    },
                                    "diskSize": {
                                        "usageRateTimeline": {
                                            "usageRate": usage
                                        }
                                    },
                                    "provisionedIops": {
                                        "usageRateTimeline": {
                                            "unit": "GiBy",
                                            "usageRateTimelineEntries": [
                                                {
                                                    "usageRate": eIOPS
                                                }
                                            ]
                                        }
                                    }

                                }
                            ],
                            "instancesRunning": {
                                    "usageRateTimeline": {
                                        "unit": "GiBs",
                                        "usageRateTimelineEntries": [
                                            {
                                            "effectiveTime": {
                                                "estimationTimeFrameOffset": "2592000s"
                                            },
                                            "usageRate": 1
                                            }
                                        ]
                                    }
                                }
                        }
                    }
                ]
            }
        }
    )
    URL = "https://cloudbilling.googleapis.com/v1beta:estimateCostScenario"

    headers = {
        "Content-Type": "application/json",
        "Accept": "*/*",
        "Connection": "keep-alive",
        "X-goog-api-key": "AIzaSyAIdlAvVrTJbEkVCuv-WZ76A6uFMRTl_ZU"
    }

    response = requests.post(URL, headers=headers, data=body)
    print(eIOPS, usage, diskType)

    

    try:
        estimate = json.loads(response.text)["costEstimationResult"]['segmentCostEstimates'][0]["segmentTotalCostEstimate"][
        "preCreditCostEstimate"]
        print ("Row -",(row) , "  Cost -" ,estimate)
        if len(estimate) == 2: #Only Units or Keys in Dict, then len(dict)==2
            if ("units") in estimate:#Only units present
                price = str(estimate["units"] + ".00")
                sheet.cell(row=row, column=5).value = round((float(price)), 2)
            else:#Only nanos present[Total 9 digits : .000 000 000]
                if len(str(estimate["nanos"])) == 9:
                    price = float("00." + str(estimate["nanos"]))
                    sheet.cell(row=row, column=5).value = (round(price, 2))
                else:#Only nanos are present[<9 digits : .000 000 00, ...]
                    price = float("0." + "0" * (9 - (len(str(estimate["nanos"])))) + str(estimate["nanos"]))
                    sheet.cell(row=row, column=5).value = (round(price, 2))
            
        else:#Both units and nanos are present in Dictionary
            if len(str(estimate["nanos"])) == 9:
                price = float(estimate["units"] + "." + str(estimate["nanos"]))
                sheet.cell(row=row, column=5).value = (round(price, 2))
            else:
                price = float(estimate["units"] + "." + "0" * (9 - (len(str(estimate["nanos"])))) + str(estimate["nanos"]))
                sheet.cell(row=row, column=5).value = (round(price, 2))
    except KeyError as e:# Not enough keys found in dictionary !
        print("Row -",(row),"\tThe Usage Rate [ ", usage ," ] , in line number : ", (row) ,", must be > 10 GiB.")
        print("\t\tThe IOPS [ ", eIOPS ," ] , in line number : ", (row) ,", must be > 2500.\n\t\tRemember : IOPS is for Extreme PD Only.")
        sheet.cell(row=row, column = 5).value = "Not Allowed"
        print (e)
workbook.save('ComputePD.xlsx')
    