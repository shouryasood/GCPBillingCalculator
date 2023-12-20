import json
import openpyxl
import requests

# Open the Excel file
workbook = openpyxl.load_workbook('Storage2.xlsx')

# Select the sheet you want to work with
sheet = workbook['Sheet1']

for row in range(2, sheet.max_row + 1):
    # for cell in row:
    storage_class = sheet.cell(row=row, column=1).value # Archieve, Coldline, Nearline , Standard 
    location = str(sheet.cell(row=row, column=2).value) # Region Name
    storage = sheet.cell(row=row, column=3).value #Storage Space in GB
    aOps = sheet.cell(row=row, column=4).value # Class A Operations/Month in Millions
    bOps = sheet.cell(row=row, column=5).value # Class B Operations/Month in Millions
    dret = sheet.cell(row=row, column=6).value # Data retrieval usage. A retrieval cost applies when data or metadata is read, copied, or rewritten 
    # redIMI = sheet.cell(row=row, column=6).value # Region Name [Inter/Multi/Intra Regional]
        # egressTraffic = sheet.cell(row=row, column=7).value # Egress Traffic for GCS in GB
    smdReg= ''#Represents particular value for Region, MultiRegion and DualRegion.
    if location in ["asia-east1","asia-east2","asia-northeast1","asia-northeast2","asia-northeast3","asia-south1","asia-south2","asia-southeast1","asia-southeast2","australia-southeast1","australia-southeast2","europe-central2","europe-north1","europe-southwest1","europe-west1","europe-west2","europe-west3","europe-west4","europe-west6","europe-west8","europe-west9","northamerica-northeast1","northamerica-northeast2","southamerica-east1","southamerica-west1","us-central1","us-east1","us-east4","us-east5","us-east7","us-south1","us-west1","us-west2","us-west3","us-west4"]:
        smdReg = '''"region"'''
    elif location in ["asia","eu","us"]:
        smdReg = '''"multiRegion"'''
    elif location in ["asia1","eur4","nam4"]:
        smdReg = '''"dualRegion"'''




    body = json.dumps({
        "costScenario": {
            "workloads": [
                {
                    "name": "storage-example",

                    "cloudStorageWorkload": 
                    {
                        "storageClass": storage_class,

                        "dataStored": {
                            "usageRateTimeline": {
                                "unit": "GiBy",
                                "usageRateTimelineEntries": [
                                    {
                                        "usageRate": storage
                                    }
                                ]
                            }
                        },
                        
                        "operationA": 
                            {
                                "usageRateTimeline": 
                                {
                                "unit": "1/s", 
                                "usageRateTimelineEntries": 
                                [
                                    {
                                    "effectiveTime": 
                                    {
                                        "estimationTimeFrameOffset": "2628000s" 
                                    },
                                    "usageRate": aOps
                                    }
                                ]
                                }    
                            },



                            "operationB": 
                            {
                                "usageRateTimeline": 
                                {
                                "unit": "1/s", 
                                "usageRateTimelineEntries": 
                                [
                                    {
                                    "effectiveTime": 
                                    {
                                        "estimationTimeFrameOffset": "2628000s" 
                                    },
                                    "usageRate": bOps 
                                    }
                                ]
                                }
                            },
                            "dataRetrieval": 
                            {
                                "usageRateTimeline": 
                                {
                                    "unit": "GiBy/s",
                                    "usageRateTimelineEntries": [
                                        {
                                        "effectiveTime": 
                                        {
                                            "estimationTimeFrameOffset": "2628000s"
                                        },
                                        "usageRate": dret
                                        }
                                    ]
                                }
                            },
                            smdReg: {
                                "name": location
                            }
                        
                    }
                }
            ],
            "scenarioConfig": 
            {
                "estimateDuration": "2628000s"
            }
        }
    }
    )

    URL = "https://cloudbilling.googleapis.com/v1beta:estimateCostScenario"

    headers = {
        "Content-Type": "application/json",
        "Accept": "*/*",
        "Connection": "keep-alive",
        "X-goog-api-key": "AIzaSyAIdlAvVrTJbEkVCuv-WZ76A6uFMRTl_ZU"
    }

    response = requests.post(URL, headers=headers, data=body)
    try:
        estimate = json.loads(response.text)["costEstimationResult"]['segmentCostEstimates'][0]["segmentTotalCostEstimate"][
        "preCreditCostEstimate"]
        print ("Row -",(row) , "  Cost -" ,estimate)
        if len(estimate) == 2: #Only Units or Keys in Dict, then len(dict)==2
            if ("units") in estimate:#Only units present
                price = str(estimate["units"] + ".00")
                sheet.cell(row=row, column=9).value = round((float(price)), 2)
            else:#Only nanos present[Total 9 digits : .000 000 000]
                if len(str(estimate["nanos"])) == 9:
                    price = float("00." + str(estimate["nanos"]))
                    sheet.cell(row=row, column=9).value = (round(price, 2))
                else:#Only nanos are present[<9 digits : .000 000 00, ...]
                    price = float("0." + "0" * (9 - (len(str(estimate["nanos"])))) + str(estimate["nanos"]))
                    sheet.cell(row=row, column=9).value = (round(price, 2))
            
        else:#Both units and nanos are present in Dictionary
            if len(str(estimate["nanos"])) == 9:
                price = float(estimate["units"] + "." + str(estimate["nanos"]))
                sheet.cell(row=row, column=9).value = (round(price, 2))
            else:
                price = float(estimate["units"] + "." + "0" * (9 - (len(str(estimate["nanos"])))) + str(estimate["nanos"]))
                sheet.cell(row=row, column=9).value = (round(price, 2))
    except KeyError:# Not enough keys found in dictionary !
        print("\nRow -",(row),"\tThe storage class [ ", storage_class ," ] , in line number : ", (row) ," is not allowed in region : ", location ,".\n")

workbook.save('Storage2.xlsx')
