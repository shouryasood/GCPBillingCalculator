import openpyxl
import requests
import json

# Open the Excel file
workbook = openpyxl.load_workbook('premiumVMDemo.xlsx')

# Select the sheet you want to work with
sheet = workbook['Sheet1']

for row in range(2, sheet.max_row + 1):
    instance_type = str(sheet.cell(row=row, column=1).value) #Machine Type of Instance
    machineSeries = str(instance_type).split("-", maxsplit=1) #Machine Series of Instance
    machineSeries = machineSeries[0]
    region = str(sheet.cell(row=row, column=2).value)
    vCpu = int(sheet.cell(row=row, column=3).value) #Number of VCPU(s)
    memGB = float(sheet.cell(row=row, column=4).value) #RAM in GB
    os = str(sheet.cell(row=row, column = 5).value) #Operating System to be used
    #Below Dictionary contains STRINGS reffering to particular OS Key as provided in Excel.
    liscdict = {   
                "FreeOrBYOL":"projects/centos-cloud/global/licenses/centos-stream-9" ,
                "P_UbuntuPro":"projects/ubuntu-os-pro-cloud/global/licenses/ubuntu-pro-2204-lts" ,
                "P_Windows":"projects/windows-cloud/global/licenses/windows-server-2022-dc" ,
                "P_RedHatEnterpriseLinux":"projects/rhel-cloud/global/licenses/rhel-9-server" ,
                "P_RedHatEnterpriseLinuxforSAP":"projects/rhel-sap-cloud/global/licenses/rhel-8-sap" ,
                "P_SLES":"projects/suse-cloud/global/licenses/sles-15" ,
                "P_SLES12forSAP":"projects/suse-sap-cloud/global/licenses/sles-sap-12" , 
                "P_SLES15forSAP":"projects/suse-sap-cloud/global/licenses/sles-sap-15" , 
                "P_SQLServerStd":"projects/windows-sql-cloud/global/licenses/sql-server-2019-standard" ,
                "P_SQLServerWeb":"projects/windows-sql-cloud/global/licenses/sql-server-2019-web" , 
                "P_SQLServerEnterprise":"projects/windows-sql-cloud/global/licenses/sql-server-2019-enterprise" 
            }
    licenses = liscdict.get(os) #Get Value from key (variable:os) to pass in JSON Request.
    body = json.dumps(
        {
            "costScenario": {
                "scenarioConfig": {
                    "estimateDuration": "2592000s"
                },
                "workloads": [
                    {
                        "name": "vm-example",
                        "computeVmWorkload": {
                            "instancesRunning": {
                                "usageRateTimeline": {
                                    "usageRateTimelineEntries": [
                                        {
                                            "usageRate": 1
                                        }
                                    ]
                                }
                            },
                            "machineType": {
                                "customMachineType": {
                                    "machineSeries": machineSeries,
                                    "virtualCpuCount": vCpu,
                                    "memorySizeGb": memGB
                                }
                            },
                            "region": region,
                            "licenses": licenses #Contains String of actual SKU
                        }
                    }
                ]
            }
        }
    )
    URL = "https://cloudbilling.googleapis.com/v1beta:estimateCostScenario"

    headers = {
        "Content-Type": "application/json",
        "Accept": "*/*",
        "Connection": "keep-alive",
        "X-goog-api-key": "AIzaSyAIdlAvVrTJbEkVCuv-WZ76A6uFMRTl_ZU"
    }

    response = requests.post(URL, headers=headers, data=body)

    

    try:
        estimate = json.loads(response.text)["costEstimationResult"]['segmentCostEstimates'][0]["segmentTotalCostEstimate"][
        "preCreditCostEstimate"]
        print ("Row -",(row) , "  Cost -" ,estimate)
        if len(estimate) == 2: #Only Units or Keys in Dict, then len(dict)==2
            if ("units") in estimate:#Only units present
                price = str(estimate["units"] + ".00")
                sheet.cell(row=row, column=6).value = round((float(price)), 2)
            else:#Only nanos present[Total 9 digits : .000 000 000]
                if len(str(estimate["nanos"])) == 9:
                    price = float("00." + str(estimate["nanos"]))
                    sheet.cell(row=row, column=6).value = (round(price, 2))
                else:#Only nanos are present[<9 digits : .000 000 00, ...]
                    price = float("0." + "0" * (9 - (len(str(estimate["nanos"])))) + str(estimate["nanos"]))
                    sheet.cell(row=row, column=6).value = (round(price, 2))
            
        else:#Both units and nanos are present in Dictionary
            if len(str(estimate["nanos"])) == 9:
                price = float(estimate["units"] + "." + str(estimate["nanos"]))
                sheet.cell(row=row, column=6).value = (round(price, 2))
            else:
                price = float(estimate["units"] + "." + "0" * (9 - (len(str(estimate["nanos"])))) + str(estimate["nanos"]))
                sheet.cell(row=row, column=6).value = (round(price, 2))
    except KeyError as e:# Not enough keys found in dictionary !
        print(e)
        sheet.cell(row=row, column = 6).value = "Not Allowed"
workbook.save('premiumVMDemo.xlsx')
    